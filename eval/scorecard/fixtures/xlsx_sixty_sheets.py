#!/usr/bin/env python3
"""Generate ``sixty-sheets.xlsx``: 60 sheets, one merged header, one wide table.

    uv run --with openpyxl python eval/scorecard/fixtures/xlsx_sixty_sheets.py [out_dir]

Sheet ``Summary`` has a merged title cell spanning three columns above a
bold header row; sheet ``Wide`` has 40 columns; the other 58 sheets are
small numeric tables so sheet ordering and naming can be checked. Values
come from a fixed arithmetic rule, never from random numbers.
"""

from __future__ import annotations

import sys
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
DEFAULT_OUT = HERE.parents[2] / "tests" / "golden" / "corpus"
FIXED_TIME = datetime(2000, 1, 1, tzinfo=UTC)
SHEETS = 60
WIDE_COLUMNS = 40
WIDE_ROWS = 12

SUMMARY_ROWS = (
    ("Region", "Units", "Revenue"),
    ("North", 120, 2400.5),
    ("South", 95, 1900.0),
    ("East", 143, 2860.25),
    ("West", 88, 1760.0),
    ("Total", 446, 8920.75),
)


def fill_summary(ws) -> None:
    ws.title = "Summary"
    ws["A1"] = "Quarterly totals by region"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")
    for c, value in enumerate(SUMMARY_ROWS[0], start=1):
        cell = ws.cell(row=2, column=c, value=value)
        cell.font = Font(bold=True)
    for r, row in enumerate(SUMMARY_ROWS[1:], start=3):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)


def fill_wide(ws) -> None:
    ws.title = "Wide"
    for c in range(1, WIDE_COLUMNS + 1):
        ws.cell(row=1, column=c, value=f"C{c:02d}").font = Font(bold=True)
    for r in range(2, WIDE_ROWS + 2):
        for c in range(1, WIDE_COLUMNS + 1):
            ws.cell(row=r, column=c, value=(r - 1) * 100 + c)
    ws.column_dimensions[get_column_letter(1)].width = 8


def fill_small(ws, index: int) -> None:
    ws.title = f"Sheet{index:02d}"
    ws.append(("Item", "Quantity", "Price"))
    for row in range(1, 4):
        ws.append((f"item-{index:02d}-{row}", index * row, round(index * row * 1.5, 2)))


def build() -> Workbook:
    wb = Workbook()
    fill_summary(wb.active)
    fill_wide(wb.create_sheet())
    for index in range(3, SHEETS + 1):
        fill_small(wb.create_sheet(), index)
    wb.properties.creator = ""
    wb.properties.lastModifiedBy = ""
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    return wb


def main(argv: list[str]) -> int:
    out_dir = Path(argv[0]) if argv else DEFAULT_OUT
    out_dir.mkdir(parents=True, exist_ok=True)
    target = out_dir / "sixty-sheets.xlsx"
    build().save(str(target))
    print(f"{target} ({target.stat().st_size} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
