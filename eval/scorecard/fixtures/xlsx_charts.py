#!/usr/bin/env python3
"""Generate ``charts.xlsx``: three sheets, each with a data table and one chart.

    uv run --with openpyxl python eval/scorecard/fixtures/xlsx_charts.py [out_dir] [render_dir]

Sheet ``Revenue`` carries a two-series clustered column chart with a title and
both axis titles; ``Temperature`` a one-series line chart with a title;
``Share`` a pie chart with no title at all. The data and titles come from
``chart_data.py``, which the truth file quotes and the derender comparison
renders. Deterministic: fixed document properties, fixed zip entry
timestamps (openpyxl stamps the wall clock otherwise; see
``chart_data.normalize_zip``). When ``render_dir`` is given the same charts
are painted to PNG there with matplotlib (needs ``--with matplotlib``).
"""

from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font

HERE = Path(__file__).resolve().parent
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))

from chart_data import BAR, CHARTS, LINE, PIE, ChartSpec, normalize_zip, render_charts  # noqa: E402

DEFAULT_OUT = HERE.parents[2] / "tests" / "golden" / "corpus"
FIXED_TIME = datetime(2000, 1, 1, tzinfo=timezone.utc)
CHART_ANCHOR = "E2"


def fill_sheet(ws, spec: ChartSpec) -> None:
    ws.title = spec.name
    for column, label in enumerate(spec.header, start=1):
        ws.cell(row=1, column=column, value=label).font = Font(bold=True)
    for row_index, row in enumerate(spec.rows, start=2):
        for column, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=column, value=value)


def add_chart(ws, spec: ChartSpec) -> None:
    last_row = len(spec.rows) + 1
    chart = {"bar": BarChart, "line": LineChart, "pie": PieChart}[spec.kind]()
    if spec.kind == "bar":
        chart.type = "col"
        chart.grouping = "clustered"
    # Header cells name the series; the first column is the category axis.
    data = Reference(ws, min_col=2, max_col=len(spec.header), min_row=1, max_row=last_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = spec.title
    if spec.kind != "pie":
        chart.x_axis.title = spec.x_title
        chart.y_axis.title = spec.y_title
        chart.x_axis.delete = False
        chart.y_axis.delete = False
    chart.width = 16
    chart.height = 8
    ws.add_chart(chart, CHART_ANCHOR)


def build() -> Workbook:
    wb = Workbook()
    sheets = [wb.active] + [wb.create_sheet() for _ in CHARTS[1:]]
    for ws, spec in zip(sheets, CHARTS):
        fill_sheet(ws, spec)
        add_chart(ws, spec)
    wb.properties.creator = ""
    wb.properties.lastModifiedBy = ""
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    return wb


def main(argv: list[str]) -> int:
    out_dir = Path(argv[0]) if argv else DEFAULT_OUT
    out_dir.mkdir(parents=True, exist_ok=True)
    target = out_dir / "charts.xlsx"
    build().save(str(target))
    normalize_zip(target)
    print(f"{target} ({target.stat().st_size} bytes)")
    if len(argv) > 1:
        for path in render_charts(Path(argv[1])):
            print(f"{path} ({path.stat().st_size} bytes)")
    return 0


if __name__ == "__main__":
    assert (BAR.series_count, LINE.series_count, PIE.series_count) == (2, 1, 1)
    sys.exit(main(sys.argv[1:]))
